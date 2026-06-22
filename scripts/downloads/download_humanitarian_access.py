#!/usr/bin/env python3
"""Download and build the latest Ethiopia humanitarian access ADM3 layer.

The source dataset is the HDX "Ethiopia - Humanitarian Access" package. The
script selects the newest XLSX edition from the package metadata, downloads the
workbook, joins the access table to the local FCV-modified ADM3 boundaries on
ADM3 p-code, and writes an AGOL-ready GeoJSON polygon layer.
"""

from __future__ import annotations

import argparse
from collections import Counter
from datetime import datetime
import json
import re
from pathlib import Path
import tempfile
from typing import Any

from openpyxl import load_workbook
import requests


DATASET_ID = "ethiopia-humanitarian-access"
CKAN_PACKAGE_URL = "https://data.humdata.org/api/3/action/package_show"
DEFAULT_ADM3 = Path("data/admin_boundaries/woredas_fcvmodified_adm3.geojson")
DEFAULT_OUTPUT = Path("data/humanitarian_access/ethiopia_humanitarian_access_adm3.geojson")
DEFAULT_WORKBOOK = Path("data/humanitarian_access/source/ethiopia_humanitarian_access_latest.xlsx")
DEFAULT_METADATA = Path("data/humanitarian_access/ethiopia_humanitarian_access_adm3_metadata.json")
USER_AGENT = "ethiopia-geo-hub-humanitarian-access/1.0"

MONTHS = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}

STATUS_LEVELS = {
    "accessible": 1,
    "some restrictions of movement": 2,
    "partially accessible": 3,
    "hard to reach": 4,
}


def clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()


def parse_timestamp(value: str | None) -> datetime:
    if not value:
        return datetime.min
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError:
        return datetime.min


def parse_month_year(text: str) -> tuple[int, int] | None:
    matches: list[tuple[int, int]] = []
    month_pattern = "|".join(sorted(MONTHS, key=len, reverse=True))
    for match in re.finditer(rf"\b({month_pattern})[\s_-]*(20\d{{2}})\b", text, re.I):
        matches.append((int(match.group(2)), MONTHS[match.group(1).lower()]))
    return max(matches) if matches else None


def period_label(period: tuple[int, int] | None) -> str:
    if not period:
        return ""
    year, month = period
    return f"{datetime(year, month, 1):%B} {year}"


def period_iso(period: tuple[int, int] | None) -> str:
    if not period:
        return ""
    year, month = period
    return f"{year:04d}-{month:02d}"


def field_name_for_access_column(header: str) -> str:
    period = parse_month_year(header)
    if period:
        year, month = period
        return f"accessibility_{datetime(year, month, 1):%B_%Y}".lower()
    name = re.sub(r"[^A-Za-z0-9_]+", "_", header).strip("_").lower()
    return name or "accessibility"


def normalize_status(value: object) -> str:
    value = clean(value)
    key = value.lower()
    labels = {
        "accessible": "Accessible",
        "some restrictions of movement": "Some restrictions of movement",
        "partially accessible": "Partially accessible",
        "hard to reach": "Hard to reach",
    }
    return labels.get(key, value)


def fetch_package(dataset_id: str, timeout: int) -> dict[str, Any]:
    response = requests.get(
        CKAN_PACKAGE_URL,
        params={"id": dataset_id},
        headers={"Accept": "application/json", "User-Agent": USER_AGENT},
        timeout=timeout,
    )
    response.raise_for_status()
    payload = response.json()
    if not payload.get("success"):
        raise RuntimeError(f"HDX package_show failed for {dataset_id}: {payload}")
    return payload["result"]


def resource_period(resource: dict[str, Any]) -> tuple[int, int] | None:
    return parse_month_year(" ".join(clean(resource.get(key)) for key in ["name", "url"]))


def choose_latest_xlsx_resource(package: dict[str, Any]) -> dict[str, Any]:
    resources = [
        resource
        for resource in package.get("resources", [])
        if clean(resource.get("format")).lower() == "xlsx" or clean(resource.get("url")).lower().endswith(".xlsx")
    ]
    if not resources:
        raise RuntimeError("No XLSX resources found in the HDX package.")

    def sort_key(resource: dict[str, Any]) -> tuple[int, int, datetime, datetime]:
        period = resource_period(resource) or (0, 0)
        return (
            period[0],
            period[1],
            parse_timestamp(resource.get("last_modified")),
            parse_timestamp(resource.get("created")),
        )

    return max(resources, key=sort_key)


def download_file(url: str, output: Path, timeout: int) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    with requests.get(url, headers={"User-Agent": USER_AGENT}, stream=True, timeout=timeout) as response:
        response.raise_for_status()
        with tempfile.NamedTemporaryFile("wb", dir=output.parent, prefix=f".{output.name}.", suffix=".tmp", delete=False) as file:
            temp_path = Path(file.name)
            for chunk in response.iter_content(chunk_size=1024 * 1024):
                if chunk:
                    file.write(chunk)
    temp_path.replace(output)


def header_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def find_access_sheet(workbook_path: Path) -> tuple[str, list[str], list[dict[str, Any]]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        try:
            raw_headers = next(rows)
        except StopIteration:
            continue
        headers = [clean(value) for value in raw_headers]
        keys = {header_key(header) for header in headers}
        if not {"admin1name", "admin2name", "admin3name"}.issubset(keys):
            continue
        if "admin3pcode" not in keys and "admin3pcod" not in keys:
            continue

        rows_out = []
        indexed_headers = [(index, header) for index, header in enumerate(headers) if header]
        for raw_row in rows:
            if not any(clean(value) for value in raw_row):
                continue
            row = {header: raw_row[index] if index < len(raw_row) else None for index, header in indexed_headers}
            if clean(row.get("Admin3PCode") or row.get("Admin3PCod")):
                rows_out.append(row)
        return sheet.title, [header for _, header in indexed_headers], rows_out
    raise RuntimeError(f"Could not find an access table sheet in {workbook_path}")


def source_pcode(row: dict[str, Any]) -> str:
    return clean(row.get("Admin3PCode") or row.get("Admin3PCod"))


def load_boundaries(path: Path) -> list[dict[str, Any]]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    features = payload.get("features") or []
    if not features:
        raise RuntimeError(f"Boundary GeoJSON has no features: {path}")
    return features


def write_json(path: Path, payload: dict[str, Any], *, compact: bool = False) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile("w", encoding="utf-8", dir=path.parent, prefix=f".{path.name}.", suffix=".tmp", delete=False) as file:
        temp_path = Path(file.name)
        if compact:
            json.dump(payload, file, ensure_ascii=False, separators=(",", ":"))
        else:
            json.dump(payload, file, ensure_ascii=False, indent=2, sort_keys=True)
        file.write("\n")
    temp_path.replace(path)


def build_layer(
    *,
    package: dict[str, Any],
    resource: dict[str, Any],
    workbook_path: Path,
    adm3_path: Path,
    output_path: Path,
    metadata_path: Path,
) -> dict[str, Any]:
    sheet_name, headers, rows = find_access_sheet(workbook_path)
    rows_by_pcode = {source_pcode(row): row for row in rows}
    access_headers = [header for header in headers if header.lower().startswith("accessibility")]
    if not access_headers:
        raise RuntimeError(f"No accessibility columns found in {workbook_path}:{sheet_name}")
    access_columns = [
        {
            "source_header": header,
            "field": field_name_for_access_column(header),
            "period": period_iso(parse_month_year(header)),
            "period_label": period_label(parse_month_year(header)),
        }
        for header in access_headers
    ]
    latest_column = max(access_columns, key=lambda column: column["period"] or "0000-00")

    features = []
    missing_source: list[str] = []
    for feature in load_boundaries(adm3_path):
        boundary_props = feature.get("properties") or {}
        pcode = clean(boundary_props.get("MergeTEXT"))
        source_row = rows_by_pcode.get(pcode)
        if not source_row:
            missing_source.append(pcode)
            source_row = {}

        latest_value = normalize_status(source_row.get(latest_column["source_header"]))
        latest_key = latest_value.lower()
        properties: dict[str, Any] = {
            "admin_level": 3,
            "adm3_pcode": pcode,
            "adm3_name": clean(boundary_props.get("admin3Name")),
            "adm3_ref_name": clean(boundary_props.get("admin3RefN")),
            "adm3_alt_name": clean(boundary_props.get("admin3AltN")),
            "adm2_name": clean(boundary_props.get("admin2Name")),
            "adm2_pcode": clean(boundary_props.get("admin2Pcod")),
            "adm1_name": clean(boundary_props.get("admin1Name")),
            "adm1_pcode": clean(boundary_props.get("admin1Pcod")),
            "source_adm3_name": clean(source_row.get("Admin3Name")),
            "source_adm3_pcode": source_pcode(source_row),
            "source_adm2_name": clean(source_row.get("Admin2Name")),
            "source_adm2_pcode": clean(source_row.get("Admin2PCode") or source_row.get("Admin2PCod")),
            "source_adm1_name": clean(source_row.get("Admin1Name")),
            "source_adm1_pcode": clean(source_row.get("Admin1PCode") or source_row.get("Admin1PCod")),
            "latest_period": latest_column["period"],
            "latest_period_label": latest_column["period_label"],
            "latest_accessibility": latest_value,
            "latest_access_level": STATUS_LEVELS.get(latest_key),
            "join_status": "matched" if source_row else "boundary_without_source",
        }
        for column in access_columns:
            properties[column["field"]] = normalize_status(source_row.get(column["source_header"]))
        features.append({"type": "Feature", "geometry": feature.get("geometry"), "properties": properties})

    output = {
        "type": "FeatureCollection",
        "name": output_path.stem,
        "crs": {"type": "name", "properties": {"name": "urn:ogc:def:crs:OGC:1.3:CRS84"}},
        "features": features,
    }
    write_json(output_path, output, compact=True)

    boundary_pcodes = {clean(feature.get("properties", {}).get("adm3_pcode")) for feature in features}
    source_without_boundary = sorted(pcode for pcode in rows_by_pcode if pcode not in boundary_pcodes)
    latest_counts = Counter(feature["properties"]["latest_accessibility"] for feature in features)
    metadata = {
        "dataset": package.get("title") or DATASET_ID,
        "dataset_id": package.get("name") or DATASET_ID,
        "package_modified": package.get("metadata_modified"),
        "source_provider": "OCHA Ethiopia",
        "source_license": package.get("license_title") or package.get("license_id"),
        "source_url": f"https://data.humdata.org/dataset/{DATASET_ID}",
        "resource": {
            "id": resource.get("id"),
            "name": resource.get("name"),
            "url": resource.get("url"),
            "format": resource.get("format"),
            "last_modified": resource.get("last_modified"),
            "created": resource.get("created"),
            "period": period_iso(resource_period(resource)),
            "period_label": period_label(resource_period(resource)),
        },
        "workbook": str(workbook_path),
        "sheet": sheet_name,
        "boundary": str(adm3_path),
        "output": str(output_path),
        "accessibility_columns": access_columns,
        "latest_column": latest_column,
        "source_row_count": len(rows),
        "source_unique_pcodes": len(rows_by_pcode),
        "feature_count": len(features),
        "matched_features": len(features) - len(missing_source),
        "boundary_without_source_count": len(missing_source),
        "boundary_without_source_pcodes": sorted(missing_source),
        "source_without_boundary_count": len(source_without_boundary),
        "source_without_boundary_pcodes": source_without_boundary,
        "latest_accessibility_counts": dict(sorted(latest_counts.items())),
    }
    write_json(metadata_path, metadata)
    return metadata


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dataset-id", default=DATASET_ID, help=f"HDX dataset id. Default: {DATASET_ID}")
    parser.add_argument("--adm3", type=Path, default=DEFAULT_ADM3, help=f"ADM3 boundary GeoJSON. Default: {DEFAULT_ADM3}")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK, help=f"Downloaded XLSX path. Default: {DEFAULT_WORKBOOK}")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help=f"Output GeoJSON path. Default: {DEFAULT_OUTPUT}")
    parser.add_argument("--metadata", type=Path, default=DEFAULT_METADATA, help=f"Output metadata JSON path. Default: {DEFAULT_METADATA}")
    parser.add_argument("--timeout", type=int, default=120, help="HTTP request timeout in seconds. Default: 120")
    parser.add_argument("--skip-download", action="store_true", help="Use the existing workbook path instead of downloading from HDX.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    package = fetch_package(args.dataset_id, args.timeout)
    resource = choose_latest_xlsx_resource(package)

    if not args.skip_download:
        download_file(resource["url"], args.workbook, args.timeout)

    metadata = build_layer(
        package=package,
        resource=resource,
        workbook_path=args.workbook,
        adm3_path=args.adm3,
        output_path=args.output,
        metadata_path=args.metadata,
    )
    print(f"Selected resource: {metadata['resource']['name']} ({metadata['resource']['period_label']})")
    print(f"Wrote {args.workbook}")
    print(f"Wrote {args.output} with {metadata['feature_count']:,} features")
    print(f"Wrote {args.metadata}")
    print(
        "Join summary: "
        f"matched={metadata['matched_features']:,}, "
        f"boundary_without_source={metadata['boundary_without_source_count']:,}, "
        f"source_without_boundary={metadata['source_without_boundary_count']:,}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
